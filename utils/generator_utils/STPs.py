import pandas as pd
import re
import os
from io import BytesIO
from openpyxl.styles import Font , Alignment, PatternFill
from dotenv import load_dotenv
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Load environment variables
load_dotenv() 

# SMTP(Simple Mail Transfer Protocol) configuration
smtp_port = 587 # Standard secure SMTP port
smtp_server = "smtp.gmail.com"  # Google SMTP Server


# Email configuration
pswd = os.getenv("Google_app_password") # Sender's Google app password
subject = "STP - " # Email subject prefix

#Function to send stp via email
def send_stp(issue_key, email_list, excel_file_content, excel_file_name, subject):
    success = True  
    
    # Iterate over each recipient's email address
    for person in email_list:
        try:
            # Email body
            body = f"""
            Hello 👋
            
            We hope you're doing well.

            We are excited to share with you the System Testing Plan (STP) for the user story {issue_key}.
            
            Thank you for your trust in our team. If you have any feedback, questions or need clarification, feel free to reach out. 🧡

            Story2Stp Team 🙌  
            """
            msg = MIMEMultipart()
            msg['From'] = "Story2STP" #Sender's email
            msg['To'] = person  #Recipients emails
            msg['Subject'] = subject #Email's subject

            # Attach a plain text body to the email message
            msg.attach(MIMEText(body, 'plain')) 

            # Create a base MIME type for attaching files (Excel file)
            attachment_package = MIMEBase('application', 'octet-stream')

            # Set the content of the attachment to the provided Excel file content
            attachment_package.set_payload(excel_file_content)
            
            # Encode the attachment content in base64 format for email transmission
            encoders.encode_base64(attachment_package)
            
            # Indicate the attatchemet filename
            attachment_package.add_header('Content-Disposition', f"attachment; filename= {excel_file_name}")
            
            # Attach the encoded attachment to the email message
            msg.attach(attachment_package)

            # Convert the entire email message into a single string format
            text = msg.as_string()

            # Establish a connection to the SMTP server for sending emails
            with smtplib.SMTP(smtp_server, smtp_port) as TIE_server:
                # Initiate a TLS connection for secure communication with the SMTP server
                TIE_server.starttls()

                # Log into the SMTP server using the sender's email address and password
                TIE_server.login(os.getenv("Google_email"), pswd)
                
                # Send the email
                TIE_server.sendmail(os.getenv("Google_email"), person, text)
        
        except Exception as e:
            success = False  

    return success  # Return success flag indicating whether email(s) were sent successfully or not
    
#Function to split the generated text into spearte test cases
def split_output_by_TestCase(output):
    test_cases = output.split("### Test Case:")[1:]  # Split the output string based on test case pattern
    return test_cases  # Return list of individual test cases

#Function to extract text between two patterns from a generated text (using regular expressions)
def text_between_patterns(pattern1, pattern2, text):
    if pattern1:
        # Find all matches for pattern1 in the text
        matches1 = list(re.finditer(pattern1, text))
        if matches1:
            # Determine the maximum end index of pattern1 matches
            max_index1 = max(match.end() for match in matches1)
        else:
            max_index1 = 0
    else:
        max_index1 = 0

    if pattern2:
        # Find all matches for pattern2 in the text
        matches2 = list(re.finditer(pattern2, text))
        if matches2:
            # Determine the minimum start index of pattern2 matches
            min_index2 = min(match.start() for match in matches2)
        else:
            min_index2 = len(text)
    else:
        min_index2 = len(text)

    # Extract text between pattern1 and pattern2
    result = text[max_index1:min_index2]

    return result.strip() if result else ""   # Return extracted text stripped

# Function to remove extra punctuation characters from the beginning or end of a sentence
def remove_extra_punctuation(sentence):
    pattern = r'^[^\w\s]+|[^\w\s]+$'    
    cleaned_sentence = re.sub(pattern, '', sentence)
    cleaned_sentence = cleaned_sentence.lstrip('\n')
    return cleaned_sentence

#Function to convert the generated output to an Excel file
def convert_output_to_excel(output):
    # Split the output string into individual test cases
    test_cases = split_output_by_TestCase(output)      

    # Initialize an empty list to store parsed test cases  
    parsed_test_cases = []

    # Iterate over each test case and extract relevent information 
    for test_case in test_cases:
        parsed_test_case = {
            "Test Case ID": remove_extra_punctuation(text_between_patterns("Test Case ID", "Test Case Description", test_case)),
            "Test Case Description": remove_extra_punctuation(text_between_patterns("Test Case Description", "Preconditions", test_case)),
            "Preconditions": remove_extra_punctuation(text_between_patterns("Preconditions", "Steps", test_case)),
            "Steps": remove_extra_punctuation(text_between_patterns("Steps", "Expected Result", test_case)),
            "Expected Result": remove_extra_punctuation(text_between_patterns("Expected Result", "", test_case))
        }
        parsed_test_cases.append(parsed_test_case)
    
    # Convert the parsed test cases into a pandas DataFrame
    df = pd.DataFrame(parsed_test_cases)

    # Create a BytesIO object to store the Excel file content
    excel_file = BytesIO()
    
    # Open a new Excel file writer using openpyxl engine
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        
        # Convert the DataFrame to an Excel sheet named 'Test Cases'
        df.to_excel(writer, index=False, sheet_name='Test Cases')

        # Get the worksheet
        worksheet = writer.sheets['Test Cases']

        # Apply basic formatting (bold font) to the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="d9e1f2", fill_type="solid")  # Specify the header color

        # Loop through each cell in the header row and apply the formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        # Define column widths for better readability
        column_widths = {
            'A': 13.29 ,  # Test Case ID
            'B': 32.14,  # Test Case Description
            'C': 45.71,  # Preconditions
            'D': 61.29,  # Steps
            'E': 32.57   # Expected Result
        }
        
        # Set the column widths according to the defined widths
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Set alignment for the cells
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Return Excel file content as a bytes string
    excel_file.seek(0)
    return excel_file.read()




